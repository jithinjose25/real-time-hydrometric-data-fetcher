import requests
from io import StringIO
import pandas as pd
from datetime import datetime
import openpyxl
import numpy as np
from datetime import datetime, timezone


USGS_station_list = ["05092000", "05054000", "05082500","05064500",	"05083500", "05102490", "05100000", "05124000", "05094000", "05101000"]

WSC_station_list = ['05NF002', '05MG001', '05MJ012', '05MH001', '05MJ001', '05MD004', '05MC001', '05MG014', 
 '05MH005', '05ME006', '05MJ003', '05ME001', '05MG013', '05ME011', '05OA007', '05MG015', 
 '05LL015', '05LE010', '05PH007', '05ME003', '05RB003', '05OA019', '05OF003', '05SA002', 
 '05OC019', '05KH015', '05KH007', '06EB004', '06EA002', '05ME005', '05KK002', '05JM015', 
 '05OB010', '05MH008', '05LJ009', '05LM006', '05OC016', '05OJ016', '05PF650', '05SD004', 
 '05NG012', '05OG005', '05MH007', '05LM001', '05RD006', '05SD005', '05SD006', '05NF007', 
 '05LG003', '05MG003', '05NF008', '06EB002', '05OJ017', '05SC002', '05OE007', '05OE015', 
 '05OC024', '05OG002', '05OG008', '05OG001', '05LK002', '05LL012', '05MD009', '05LM005', 
 '05MF020', '05RD005', '05SB006', '05SG001', '05SD001', '05SA003', '05LH001', '05OF024', 
 '05MF001', '05MF018', '05MH006', '05OA006', '05OD028', '05RA001', '05OE006', '05OE010', 
 '05NG020', '05LJ019', '05OF025', '05OF020', '05LJ025', '05OB021', '05OJ008', '05OF026', 
 '05LG004', '05PF070', '05MG004', '05LJ005', '05MJ013', '05LD001', '05OA010', '05OB023', 
 '05OB001', '05OB019', '05OB007', '05OA001', '05NG003', '05NG024', '05NG007', '05LL019', 
 '05MJ006', '06EA012', '05JM001', '05OE001', '05OE004', '05LC001', '05LC004', '05OC021', 
 '05OC026', '05OJ022', '05OC001', '05OJ015', '05OC006', '05OJ005', '05OC028', '05OC009', 
 '05OC020', '05OC029', '05OC010', '05OC012', '05OJ021', '05OC022', '05LE005', '05OB027', 
 '05MF008', '05OD004', '05OD001', '05OF009', '05KJ001', '05ME009', '05OE011', '05OH007', 
 '05OH009', '05OF014', '05MD005', '05MD007', '05ME010', '05ME007', '05OB016', '05NF001', 
 '05NG021', '05NG001', '05MD010', '05MJ004', '05LE001', '05LE006', '05OF018', '05OE009', 
 '05LJ007', '05LJ010', '05LJ012', '05OD034', '05LH005', '05PG811', '05PH003', '05LL002', 
 '05LL005', '05MB003', '05PG001', '05NG023', '05LJ045', '05PF062', '05LE004']


def retrieve_and_process_USGS_water_services_data(start_date, end_date):

    usgs_url_endpoint = r"https://waterservices.usgs.gov/nwis/iv/"

    params = {
        "sites" : ",".join(USGS_station_list),
        "parameterCd" : "00060,00065",
        "startDT" : start_date,
        "endDT" : end_date,
        "format": "json"
    }


    response = requests.get(usgs_url_endpoint, params=params)
    data = response.json()

    raw_discharge_records_USGS = []
    raw_waterlevel_records_USGS = []


    for data in data["value"]["timeSeries"]:
        if (data["variable"]["variableCode"][0]["value"]) == "00060":
            for value in (data["values"][0]["value"]):
                discharge_value = (value["value"])
                Timestamp_local = (value["dateTime"])
                Timestamp_local_dt_obj = datetime.fromisoformat(Timestamp_local)
                Timestamp_utc_dt_obj = Timestamp_local_dt_obj.astimezone(timezone.utc)
                station_ID = data["sourceInfo"]["siteCode"][0]["value"]
                row = {
                    "Timestamp Local": Timestamp_utc_dt_obj,
                    "Station ID": station_ID,
                    "Discharge": discharge_value                  
                }

                raw_discharge_records_USGS.append(row)

        elif (data["variable"]["variableCode"][0]["value"]) == "00065":
            for value in (data["values"][0]["value"]):
                waterlevel_value = (value["value"])
                Timestamp_local = (value["dateTime"])
                Timestamp_local_dt_obj = datetime.fromisoformat(Timestamp_local)
                Timestamp_utc_dt_obj = Timestamp_local_dt_obj.astimezone(timezone.utc)
                station_ID = data["sourceInfo"]["siteCode"][0]["value"]
                row = {
                    "Timestamp Local": Timestamp_utc_dt_obj,
                    "Station ID": station_ID,
                    "WaterLevel": waterlevel_value                  
                }

                raw_waterlevel_records_USGS.append(row)

    raw_discharge_records_USGS_df = pd.DataFrame(raw_discharge_records_USGS)
    raw_waterlevel_records_USGS_df = pd.DataFrame(raw_waterlevel_records_USGS)
    raw_discharge_records_USGS_df["Date of Observation"] = raw_discharge_records_USGS_df["Timestamp Local"].dt.date
    raw_discharge_records_USGS_df["Discharge"] = pd.to_numeric(raw_discharge_records_USGS_df["Discharge"])
    raw_waterlevel_records_USGS_df["Date of Observation"] = raw_waterlevel_records_USGS_df["Timestamp Local"].dt.date
    raw_waterlevel_records_USGS_df["WaterLevel"] = pd.to_numeric(raw_waterlevel_records_USGS_df["WaterLevel"])

    raw_waterlevel_records_USGS_df.loc[raw_waterlevel_records_USGS_df["WaterLevel"] == -999999, "WaterLevel"] = np.nan
    raw_discharge_records_USGS_df.loc[raw_discharge_records_USGS_df["Discharge"] == -999999, "Discharge"] = np.nan

    daily_discharge_records_USGS_df = raw_discharge_records_USGS_df.groupby(["Station ID","Date of Observation"]).agg({"Discharge" : "mean"}).reset_index()
    daily_discharge_records_USGS_df.rename(columns = { "Discharge" : "Daily_Averaged_Discharge"}, inplace=True)
    daily_waterlevel_records_USGS_df = raw_waterlevel_records_USGS_df.groupby(["Station ID","Date of Observation"]).agg({"WaterLevel" : "mean"}).reset_index()
    daily_waterlevel_records_USGS_df.rename(columns = {"WaterLevel" : "Daily_Averaged_WaterLevel"}, inplace=True)
    daily_discharge_records_USGS_df = daily_discharge_records_USGS_df[["Date of Observation", "Station ID", "Daily_Averaged_Discharge"]]
    daily_waterlevel_records_USGS_df = daily_waterlevel_records_USGS_df[["Date of Observation", "Station ID", "Daily_Averaged_WaterLevel"]]

    return daily_discharge_records_USGS_df, daily_waterlevel_records_USGS_df



def retrieve_and_process_WSC_eccc_data(start_date, end_date):

    eccc_url_endpoint = "https://api.weather.gc.ca/collections/hydrometric-realtime/items"

    formatted_datetime_str = f"{start_date}/{end_date}"
    eccc_raw_records =[]

    for station in WSC_station_list:

        params = {
            "STATION_NUMBER" : station,
            "datetime" : formatted_datetime_str,
            "f": "json",
            "limit":100000
        }

        response = requests.get(eccc_url_endpoint, params = params)
        data = response.json()
        features = data["features"]



        for feature in features:
            station_waterlevel = feature["properties"].get("LEVEL")
            station_discharge = feature["properties"].get("DISCHARGE")
            station_id = (feature["properties"].get("STATION_NUMBER"))
            date_of_observation = feature["properties"].get("DATETIME")
            row = {
                "Station ID": station_id,
                "DISCHARGE": station_discharge,
                "WATERLEVEL": station_waterlevel,
                "TIME STAMP": date_of_observation
            }
            eccc_raw_records.append(row)


    eccc_raw_records_df = pd.DataFrame(eccc_raw_records)
    eccc_raw_records_df["TIME STAMP"] = pd.to_datetime(eccc_raw_records_df["TIME STAMP"])
    eccc_raw_records_df["DATE"] = eccc_raw_records_df["TIME STAMP"].dt.date

    eccc_raw_records_df.loc[eccc_raw_records_df["WATERLEVEL"] == -999.0, "WATERLEVEL"] = np.nan
    eccc_raw_records_df.loc[eccc_raw_records_df["DISCHARGE"] == -999.0, "DISCHARGE"] = np.nan

    eccc_raw_records_df.to_csv(r"C:\Users\jithinjose\Downloads\test.csv", sep=',')

    eccc_daily_averaged_df = eccc_raw_records_df.groupby(["Station ID","DATE"]).agg({"DISCHARGE": "mean", "WATERLEVEL": "mean"}).reset_index()
    eccc_daily_averaged_df.rename(columns={"DATE":"Date of Observation", "DISCHARGE":"Daily_Averaged_Discharge", "WATERLEVEL":"Daily_Averaged_WaterLevel"}, inplace=True)
    df_daily_discharge_eccc = eccc_daily_averaged_df[["Date of Observation", "Station ID", "Daily_Averaged_Discharge"]]
    df_daily_waterlevel_eccc = eccc_daily_averaged_df[["Date of Observation", "Station ID", "Daily_Averaged_WaterLevel"]]

    print(df_daily_waterlevel_eccc.head())
    return df_daily_discharge_eccc, df_daily_waterlevel_eccc



def retrieve_mapping_from_masterexcel_discharge_records(workbook_path):

    wb = openpyxl.load_workbook(workbook_path)
    ws_discharge = wb["FLOW"]

    station_id_cell_index_mapping = {} 

    for cell in ws_discharge[2][1:]:
        if cell.value:
            station_id_cell_index_mapping[cell.value] = cell.column

    date_cell_index_mapping = {}

    for row in ws_discharge.iter_rows(min_row = 4 ,min_col = 1, max_col= 1):
        date_cell = row[0]
        if date_cell.value:
            date_cell_index_mapping[date_cell.value] = date_cell.row

    return station_id_cell_index_mapping, date_cell_index_mapping, ws_discharge


def write_discharge_records_to_masterexcel(station_index_mapping, date_index_mapping, pivoted_discharge_records_df,ws_discharge, workbook_path):

    for index_date, row_data in pivoted_discharge_records_df.iterrows():
        if index_date in date_index_mapping:
            row_write_index = date_index_mapping[index_date]
            for station_id,discharge in row_data.items():
                if station_id in station_index_mapping:
                    col_write_index = station_index_mapping[station_id]
                    ws_discharge.cell(row = row_write_index, column = col_write_index, value = discharge)

    wb = ws_discharge.parent
    wb.save(workbook_path)

def retrieve_mapping_from_masterexcel_waterlevel_records(workbook_path):
                
    wb = openpyxl.load_workbook(workbook_path)
    ws_waterlevel = wb["LEVEL"]

    station_id_cell_index_mapping = {}

    for cell in ws_waterlevel[2][1:]:
        if cell.value:
            station_id_cell_index_mapping[cell.value] = cell.column

    date_cell_index_mapping = {}

    for row in ws_waterlevel.iter_rows(min_row = 4, min_col= 1, max_col = 1):
        date_cell = row[0]
        if date_cell.value:
            date_cell_index_mapping[date_cell.value] = date_cell.row

    return station_id_cell_index_mapping, date_cell_index_mapping, ws_waterlevel



def write_waterlevel_records_to_masterexcel(station_index_mapping, date_index_mapping, pivoted_waterlevel_records_df, ws_waterlevel, workbook_path):

    for index_date, row_data in pivoted_waterlevel_records_df.iterrows():
        if index_date in date_index_mapping:
            row_write_index = date_index_mapping[index_date]
            for station_id, waterlevel in row_data.items():
                if station_id in station_index_mapping:
                    col_write_index = station_index_mapping[station_id]
                    ws_waterlevel.cell(row = row_write_index, column = col_write_index, value = waterlevel)
        

    wb = ws_waterlevel.parent
    wb.save(workbook_path)


def main():
    start_date = str(input("Enter start date (YYYY-MM-DD): "))
    end_date = str(input("Enter end date (YYYY-MM-DD): "))

    workbook_path = str(input("Enter the full file path for the master excel workbook wherein the data gets appended: "))

    USGS_discharge_records_df,USGS_waterlevel_records_df = retrieve_and_process_USGS_water_services_data(start_date, end_date) 
    WSC_discharge_records_df, WSC_waterlevel_records_df = retrieve_and_process_WSC_eccc_data(start_date, end_date)
    combined_discharge_records_df = pd.concat([USGS_discharge_records_df,WSC_discharge_records_df], ignore_index=True)
    combined_discharge_records_df["Date of Observation"] = pd.to_datetime(combined_discharge_records_df["Date of Observation"]).dt.floor("D") 

    combined_waterlevel_records_df = pd.concat([USGS_waterlevel_records_df,WSC_waterlevel_records_df], ignore_index=True)
    combined_waterlevel_records_df["Date of Observation"] = pd.to_datetime(combined_waterlevel_records_df["Date of Observation"]).dt.floor("D") 
    print(combined_waterlevel_records_df.head())

    pivoted_discharge_records_df = combined_discharge_records_df.pivot(index = "Date of Observation", columns = "Station ID", values = "Daily_Averaged_Discharge")
    pivoted_discharge_records_df = pivoted_discharge_records_df.replace({np.nan: "NaN"})

    pivoted_waterlevel_records_df = combined_waterlevel_records_df.pivot(index = "Date of Observation", columns = "Station ID", values = "Daily_Averaged_WaterLevel")
    pivoted_waterlevel_records_df = pivoted_waterlevel_records_df.replace({np.nan: "NaN"})
    
    station_index_mapping_discharge_rec, date_index_mapping_discharge_rec, ws_discharge = retrieve_mapping_from_masterexcel_discharge_records(workbook_path)

    write_discharge_records_to_masterexcel(station_index_mapping_discharge_rec, date_index_mapping_discharge_rec, pivoted_discharge_records_df,ws_discharge, workbook_path)

    station_index_mapping_waterlevel_rec, date_index_mapping_waterlevel_rec, ws_waterlevel = retrieve_mapping_from_masterexcel_waterlevel_records(workbook_path)

    write_waterlevel_records_to_masterexcel(station_index_mapping_waterlevel_rec, date_index_mapping_waterlevel_rec, pivoted_waterlevel_records_df, ws_waterlevel, workbook_path)


if __name__ == "__main__":
    main()
